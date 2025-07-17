from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Create a workbook and sheets
wb = Workbook()
ws = wb.active
ws.title = "Inventory"

# Headers
headers = [
    "Item ID", "Item Name", "Category", "Quantity In Stock",
    "Reorder Level", "Unit Price", "Total Value", "Low Stock Alert"
]
ws.append(headers)

# Sample data
sample_data = [
    [1001, "Pen", "Stationery", 150, 50, 0.5],
    [1002, "Notebook", "Stationery", 75, 30, 1.2],
    [1003, "Stapler", "Stationery", 25, 40, 3.5],
    [1004, "Marker", "Stationery", 10, 20, 1.0]
]

for row in sample_data:
    total_value = row[3] * row[5]
    current_row = len(ws['A']) + 1
    low_stock_formula = f'=IF(D{current_row}<E{current_row},"LOW","OK")'
    ws.append(row + [total_value, low_stock_formula])

# Format header
header_font = Font(bold=True)
for cell in ws[1]:
    cell.font = header_font
    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

# Add table
table = Table(displayName="InventoryTable", ref=f"A1:H{ws.max_row}")
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# Add Stock Movement Sheet
log_ws = wb.create_sheet(title="Stock_Log")
log_headers = ["Date", "Item ID", "Movement Type", "Quantity", "Notes"]
log_ws.append(log_headers)

# Format log header
for cell in log_ws[1]:
    cell.font = header_font
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# Save the file
wb.save("Inventory_Management_Template.xlsx")
